"""
Build Todo_Directo_Calculator_WebApp.xlsx
-- Designed for: immediate rep use AND future porting to a web app.
-- All pricing/commission data lives in a structured 'DATA' sheet.
-- Quoter and Commission sheets reference DATA via named ranges.
-- Sheet 4 exports a flat JSON-friendly summary table ready for API use.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
OUT = os.path.join(os.path.dirname(__file__), "..", "Todo_Directo_Calculator_WebApp.xlsx")

# ── Palette ────────────────────────────────────────────────────────
C = {
    "navy":    "1A1A2E", "blue":   "0F3460", "red":    "E94560",
    "gold":    "F5A623", "white":  "FFFFFF", "light":  "F0F4FF",
    "input":   "EBF5FB", "green":  "D4EDDA", "gray":   "F8F9FA",
    "dkgray":  "555555", "black":  "000000", "blue2":  "16213E",
}

def fill(hex_): return PatternFill("solid", start_color=hex_)
def ft(bold=False, color="000000", size=10, italic=False, name="Calibri"):
    return Font(name=name, bold=bold, color=color, size=size, italic=italic)
def border():
    s = Side(style="thin", color="C5D8F0")
    return Border(left=s, right=s, top=s, bottom=s)
def thick_border():
    s = Side(style="medium", color="0F3460")
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def w(ws, col_letter, width): ws.column_dimensions[col_letter].width = width
def h(ws, row, height):       ws.row_dimensions[row].height = height

def hdr_cell(ws, ref, text, bg="1A1A2E", fg="FFFFFF", size=11, bold=True, halign="center", wrap=False):
    c = ws[ref]
    c.value = text; c.font = ft(bold=bold, color=fg, size=size)
    c.fill = fill(bg); c.alignment = align(halign, wrap=wrap); c.border = border()

def inp_cell(ws, ref, val=None, fmt=None):
    c = ws[ref]
    c.value = val; c.font = ft(color="0000FF")
    c.fill = fill(C["input"]); c.alignment = align(); c.border = border()
    if fmt: c.number_format = fmt

def fml_cell(ws, ref, formula, fmt=None, bold=False, fg="000000", bg="FFFFFF"):
    c = ws[ref]
    c.value = formula; c.font = ft(bold=bold, color=fg)
    c.fill = fill(bg); c.alignment = align(); c.border = border()
    if fmt: c.number_format = fmt

def lbl_cell(ws, ref, text, bg="FFFFFF", fg="000000", bold=False, halign="left", size=10, italic=False):
    c = ws[ref]
    c.value = text; c.font = ft(bold=bold, color=fg, size=size, italic=italic)
    c.fill = fill(bg); c.alignment = align(halign); c.border = border()

def sec_hdr(ws, row, c1, c2, text, bg="0F3460"):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cc = ws.cell(row=row, column=c1)
    cc.value = text; cc.font = ft(bold=True, color="FFFFFF", size=11)
    cc.fill = fill(bg); cc.alignment = align("center"); cc.border = border()

# All commission / pricing data – single source of truth
PRODUCTS = [
    # id, service, plan, client_setup, client_monthly, rep_upfront, rep_residual_pct, rep_upfront_annual, category
    ("ALT_FREE",  "AlToque App",        "Chivo (Gratis)",              0,   0,     0,      0,    0,      "altoque"),
    ("ALT_PRO",   "AlToque App",        "Pro",                         0,   14.99, 22.49,  0,    37.50,  "altoque"),
    ("ALT_PREM",  "AlToque App",        "Premium",                     0,   29.99, 44.98,  0,    75.00,  "altoque"),
    ("WEB_CHV",   "Desarrollo Web",     "Plan Chivo",                  99,  25,    39.60,  0.10, 0,      "web"),
    ("WEB_PRO",   "Desarrollo Web",     "Plan Pro",                    149, 49,    59.60,  0.10, 0,      "web"),
    ("WEB_PREM",  "Desarrollo Web",     "Plan Premium",                299, 89,    119.60, 0.10, 0,      "web"),
    ("VAC_A",     "Vacation Rentals",   "Modelo A – Direct Booking",   300, 0,     300,    0,    0,      "vacation"),
    ("VAC_B",     "Vacation Rentals",   "Modelo B – Full Management",  600, 0,     600,    0,    0,      "vacation"),
    ("AUT_BAS",   "Automatización",     "Básica",                      0,   200,   100,    0.15, 0,      "automation"),
    ("AUT_PRO",   "Automatización",     "Pro",                         0,   400,   200,    0.15, 0,      "automation"),
    ("CHAT_BAS",  "Chatbot IA",         "Básico",                      0,   150,   75,     0.15, 0,      "chatbot"),
    ("CHAT_PRO",  "Chatbot IA",         "Pro",                         0,   300,   150,    0.15, 0,      "chatbot"),
]

BONUSES = [
    ("weekly",    "Bono Semanal",    5,  50),
    ("monthly",   "Bono Mensual",    20, 200),
    ("quarterly", "Bono Trimestral", 60, 500),
]

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
#  SHEET 1: DATA  (source of truth – do not rename; webapp reads this)
# ════════════════════════════════════════════════════════════════════
ws_data = wb.active
ws_data.title = "DATA"
ws_data.sheet_view.showGridLines = False

for col, width in zip("ABCDEFGHIJ", [14,22,26,16,16,16,16,16,12,14]):
    w(ws_data, col, width)

# Title
ws_data.merge_cells("A1:J1")
hdr_cell(ws_data, "A1", "TODO DIRECTO — TABLA MAESTRA DE PRECIOS Y COMISIONES (Fuente de Datos para Web App)",
         bg=C["navy"], size=12)
h(ws_data, 1, 30)

ws_data.merge_cells("A2:J2")
hdr_cell(ws_data, "A2", "⚠️  NO modificar IDs ni nombres de columna  —  Esta hoja es la fuente para el Cotizador y para la integración web.",
         bg=C["blue2"], size=9, fg="AABBCC")
h(ws_data, 2, 18)
h(ws_data, 3, 6)

# Column headers
DATA_HDR_ROW = 4
DATA_COLS = ["product_id","service","plan","client_setup_usd","client_monthly_usd",
             "rep_upfront_usd","rep_residual_pct","rep_annual_usd","category","active"]
for ci, col_name in enumerate(DATA_COLS, 1):
    c = ws_data.cell(row=DATA_HDR_ROW, column=ci)
    c.value = col_name; c.font = ft(bold=True, color="FFFFFF", size=9)
    c.fill = fill(C["red"]); c.alignment = align(wrap=True); c.border = border()
h(ws_data, DATA_HDR_ROW, 30)

DATA_START = 5
for i, row in enumerate(PRODUCTS):
    r = DATA_START + i
    h(ws_data, r, 20)
    bg = C["light"] if i % 2 == 0 else C["white"]
    pid, svc, plan, c_setup, c_mo, rep_up, rep_res_pct, rep_ann, cat = row
    vals = [pid, svc, plan, c_setup, c_mo, rep_up, rep_res_pct, rep_ann, cat, "YES"]
    fmts = [None, None, None, '$#,##0.00', '$#,##0.00', '$#,##0.00', '0%', '$#,##0.00', None, None]
    for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
        cc = ws_data.cell(row=r, column=ci)
        cc.value = v; cc.font = ft(color="0000FF")
        cc.fill = fill(C["input"]); cc.alignment = align(); cc.border = border()
        if fmt: cc.number_format = fmt

DATA_END = DATA_START + len(PRODUCTS) - 1
h(ws_data, DATA_END + 1, 6)

# Bonus table
BONUS_HDR_ROW = DATA_END + 2
ws_data.merge_cells(f"A{BONUS_HDR_ROW}:D{BONUS_HDR_ROW}")
hdr_cell(ws_data, f"A{BONUS_HDR_ROW}", "TABLA DE BONOS POR VOLUMEN", bg=C["blue"])
h(ws_data, BONUS_HDR_ROW, 22)

for ci, bh in enumerate(["bonus_id","nombre","target_deals","prize_usd"], 1):
    c = ws_data.cell(row=BONUS_HDR_ROW+1, column=ci)
    c.value = bh; c.font = ft(bold=True, color="FFFFFF", size=9)
    c.fill = fill(C["red"]); c.alignment = align(); c.border = border()
h(ws_data, BONUS_HDR_ROW+1, 22)

for i, (bid, bname, btarget, bprize) in enumerate(BONUSES):
    r = BONUS_HDR_ROW + 2 + i
    h(ws_data, r, 20)
    bg = C["light"] if i % 2 == 0 else C["white"]
    for ci, v in enumerate([bid, bname, btarget, bprize], 1):
        cc = ws_data.cell(row=r, column=ci)
        cc.value = v; cc.font = ft(color="0000FF")
        cc.fill = fill(C["input"]); cc.alignment = align(); cc.border = border()
        if ci == 4: cc.number_format = '$#,##0.00'

# Named ranges for cross-sheet use (note: define via wb.defined_names)
# We'll use absolute refs in formulas instead

# ════════════════════════════════════════════════════════════════════
#  SHEET 2: COTIZADOR (Client Quote)
# ════════════════════════════════════════════════════════════════════
ws_q = wb.create_sheet("Cotizador de Clientes")
ws_q.sheet_view.showGridLines = False

for col, width in zip("ABCDEFGHIJ", [14,28,14,14,14,14,16,16,16,16]):
    w(ws_q, col, width)

# Title banner
ws_q.merge_cells("A1:J1")
hdr_cell(ws_q, "A1", "TODO DIRECTO — COTIZADOR DE CLIENTE", bg=C["navy"], size=14)
h(ws_q, 1, 36)
ws_q.merge_cells("A2:J2")
hdr_cell(ws_q, "A2", "Herramienta de Cotización | Completa la sección de información y escribe SÍ en la columna 'Incluir' para agregar el servicio.", bg=C["blue2"], size=9, fg="AABBCC")
h(ws_q, 2, 18); h(ws_q, 3, 8)

# Client info
sec_hdr(ws_q, 4, 1, 10, "INFORMACIÓN DEL CLIENTE", bg=C["blue2"])
h(ws_q, 4, 24)
for r in [5, 6, 7]: h(ws_q, r, 22)

lbl_cell(ws_q, "A5", "Nombre del Cliente:", bold=True, halign="right")
ws_q.merge_cells("B5:D5"); inp_cell(ws_q, "B5")
lbl_cell(ws_q, "E5", "Representante:", bold=True, halign="right")
ws_q.merge_cells("F5:H5"); inp_cell(ws_q, "F5")

lbl_cell(ws_q, "A6", "Nombre del Negocio:", bold=True, halign="right")
ws_q.merge_cells("B6:D6"); inp_cell(ws_q, "B6")
lbl_cell(ws_q, "E6", "Fecha:", bold=True, halign="right")
ws_q.merge_cells("F6:H6"); inp_cell(ws_q, "F6", fmt="DD/MM/YYYY")

lbl_cell(ws_q, "A7", "Tipo de Negocio:", bold=True, halign="right")
ws_q.merge_cells("B7:D7"); inp_cell(ws_q, "B7")
lbl_cell(ws_q, "E7", "Ciudad:", bold=True, halign="right")
ws_q.merge_cells("F7:H7"); inp_cell(ws_q, "F7")

h(ws_q, 8, 8)

# Service table header
sec_hdr(ws_q, 9, 1, 10, "SERVICIOS Y COTIZACIÓN", bg=C["blue"])
h(ws_q, 9, 24)

q_cols = ["ID\n(Web Ref)", "Servicio", "Plan/Paquete", "Incluir?\n(SÍ/NO)",
          "Instalación\nCliente ($)", "Mensualidad\nCliente ($)",
          "Anual\n(x10 meses)", "Ahorro\nAnual ($)", "Total Mensual\n(Setup+1mo)", "Total Anual\n(Setup+10mo)"]
for ci, h_txt in enumerate(q_cols, 1):
    cc = ws_q.cell(row=10, column=ci)
    cc.value = h_txt; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["red"]); cc.alignment = align(wrap=True); cc.border = border()
h(ws_q, 10, 38)

# Reference DATA rows
QSTART = 11
for i, prod in enumerate(PRODUCTS):
    r = QSTART + i
    h(ws_q, r, 22)
    d_row = DATA_START + i  # row number in DATA sheet
    bg = C["light"] if i % 2 == 0 else C["white"]

    # A: product_id (pulled from DATA)
    fml_cell(ws_q, f"A{r}", f"=DATA!A{d_row}", bg=bg, fg="008000")
    # B: service (pulled from DATA)
    fml_cell(ws_q, f"B{r}", f"=DATA!B{d_row}", bg=bg, fg="008000")
    ws_q[f"B{r}"].alignment = align("left")
    # C: plan (pulled from DATA)
    fml_cell(ws_q, f"C{r}", f"=DATA!C{d_row}", bg=bg, fg="008000")
    ws_q[f"C{r}"].alignment = align("left")
    # D: include input (blue)
    inp_cell(ws_q, f"D{r}", "NO")
    # E: client setup (from DATA, blue input if want to override)
    fml_cell(ws_q, f"E{r}", f"=DATA!D{d_row}", fmt='$#,##0.00;($#,##0.00);"-"', bg=bg, fg="008000")
    # F: client monthly (from DATA)
    fml_cell(ws_q, f"F{r}", f"=DATA!E{d_row}", fmt='$#,##0.00;($#,##0.00);"-"', bg=bg, fg="008000")
    # G: Annual option price (10 months)
    fml_cell(ws_q, f"G{r}",
        f'=IF(F{r}>0,F{r}*10,"-")',
        fmt='$#,##0.00;($#,##0.00);"-"', bg=bg)
    # H: Savings vs paying monthly for 12 months
    fml_cell(ws_q, f"H{r}",
        f'=IF(UPPER(D{r})="SÍ",IF(ISNUMBER(G{r}),F{r}*12-(E{r}+G{r}),0),"-")',
        fmt='$#,##0.00;($#,##0.00);"-"', bg=bg, fg="008000")
    # I: Total monthly (client pays: setup + 1 month)
    fml_cell(ws_q, f"I{r}",
        f'=IF(UPPER(D{r})="SÍ",E{r}+F{r},"-")',
        fmt='$#,##0.00;($#,##0.00);"-"', bg=bg)
    # J: Total annual (setup + 10 months)
    fml_cell(ws_q, f"J{r}",
        f'=IF(UPPER(D{r})="SÍ",IF(ISNUMBER(G{r}),E{r}+G{r},E{r}),"-")',
        fmt='$#,##0.00;($#,##0.00);"-"', bg=bg)

QEND = QSTART + len(PRODUCTS) - 1
h(ws_q, QEND + 1, 8)

# Totals
SUM_R = QEND + 2
sec_hdr(ws_q, SUM_R, 1, 10, "RESUMEN DE COTIZACIÓN PARA EL CLIENTE", bg=C["navy"])
h(ws_q, SUM_R, 26)

totals = [
    (SUM_R+1, "Cuota Total de Instalación (seleccionados):",
     f'=SUMIF(D{QSTART}:D{QEND},"SÍ",E{QSTART}:E{QEND})', C["light"]),
    (SUM_R+2, "Mensualidad Total del Cliente:",
     f'=SUMIF(D{QSTART}:D{QEND},"SÍ",F{QSTART}:F{QEND})', C["white"]),
    (SUM_R+3, "Total Opción Anual (Setup + 10 meses):",
     f'=SUMIF(D{QSTART}:D{QEND},"SÍ",J{QSTART}:J{QEND})', C["light"]),
    (SUM_R+4, "Ahorro Total del Cliente (Anual vs Mensual×12):",
     f'=SUMIF(D{QSTART}:D{QEND},"SÍ",H{QSTART}:H{QEND})', C["green"]),
]
for row_n, label_t, formula, bg in totals:
    h(ws_q, row_n, 22)
    ws_q.merge_cells(f"A{row_n}:H{row_n}")
    lbl_cell(ws_q, f"A{row_n}", label_t, bold=True, halign="right", bg=bg)
    ws_q.merge_cells(f"I{row_n}:J{row_n}")
    fml_cell(ws_q, f"I{row_n}", formula, fmt='$#,##0.00', bold=True, bg=bg, fg="008000")

# Instruction
NOTE_R = SUM_R + 6
h(ws_q, NOTE_R, 32)
ws_q.merge_cells(f"A{NOTE_R}:J{NOTE_R}")
c_note = ws_q[f"A{NOTE_R}"]
c_note.value = ("📌 INSTRUCCIONES: Escribe 'SÍ' en columna D para incluir ese servicio. "
                "Columnas A–C y E–F se llenan automáticamente desde la hoja DATA. "
                "Para cambiar un precio puntual, edita directamente en la hoja DATA. "
                "Las columnas I y J se calculan solas.")
c_note.font = Font(name="Calibri", size=9, italic=True, color="555555")
c_note.fill = fill("FFF8DC"); c_note.alignment = align("left", wrap=True); c_note.border = border()

ws_q.freeze_panes = "A11"

# ════════════════════════════════════════════════════════════════════
#  SHEET 3: COMMISSION TRACKER
# ════════════════════════════════════════════════════════════════════
ws_c = wb.create_sheet("Calculadora de Comisiones")
ws_c.sheet_view.showGridLines = False

for col, width in zip("ABCDEFGHIJ", [14,26,22,16,16,16,16,16,16,4]):
    w(ws_c, col, width)

ws_c.merge_cells("A1:I1")
hdr_cell(ws_c, "A1", "TODO DIRECTO — CALCULADORA DE COMISIONES DEL REPRESENTANTE", bg=C["navy"], size=13)
h(ws_c, 1, 34)
ws_c.merge_cells("A2:I2")
hdr_cell(ws_c, "A2", "Comisiones se calculan desde la hoja DATA. Ingresa los cierres en la tabla de registro abajo.", bg=C["blue2"], size=9, fg="AABBCC")
h(ws_c, 2, 18); h(ws_c, 3, 8)

# Rep info
sec_hdr(ws_c, 4, 1, 9, "INFORMACIÓN DEL REPRESENTANTE", bg=C["blue2"])
h(ws_c, 4, 24); h(ws_c, 5, 22); h(ws_c, 6, 22)

lbl_cell(ws_c, "A5", "Representante:", bold=True, halign="right")
ws_c.merge_cells("B5:D5"); inp_cell(ws_c, "B5")
lbl_cell(ws_c, "E5", "Mes de Seguimiento:", bold=True, halign="right")
ws_c.merge_cells("F5:H5"); inp_cell(ws_c, "F5", fmt="MMMM YYYY")

lbl_cell(ws_c, "A6", "Meta de Cierres del Mes:", bold=True, halign="right")
ws_c.merge_cells("B6:D6"); inp_cell(ws_c, "B6", 20)
lbl_cell(ws_c, "E6", "Progress hacia Meta:", bold=True, halign="right")
ws_c.merge_cells("F6:H6")
fml_cell(ws_c, "F6",
    '=IFERROR(COUNTA(B20:B39)/B6,"—")',
    fmt="0.0%", bold=True, fg="008000", bg=C["light"])

h(ws_c, 7, 8)

# Commission rate reference (auto-pulled from DATA)
sec_hdr(ws_c, 8, 1, 9, "TASAS DE COMISIÓN (Referencia automática desde DATA — no editar aquí)", bg=C["blue"])
h(ws_c, 8, 24)

rate_headers = ["product_id", "Servicio", "Plan", "Cliente\nMensual ($)", "Tu Comisión\nInicial ($)",
                "% Residual\nMensual", "Tu Anual\nUpfront ($)", "Tu Residual\n(1er mes est.)"]
for ci, rh in enumerate(rate_headers, 1):
    cc = ws_c.cell(row=9, column=ci)
    cc.value = rh; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["red"]); cc.alignment = align(wrap=True); cc.border = border()
h(ws_c, 9, 36)

RATE_START = 10
for i, prod in enumerate(PRODUCTS):
    r = RATE_START + i
    h(ws_c, r, 20)
    d_row = DATA_START + i
    bg = C["light"] if i % 2 == 0 else C["white"]
    # Pull all values from DATA
    for ci, col in enumerate(["A","B","C","E","F","G","H"], 1):
        fml_cell(ws_c, f"{get_column_letter(ci)}{r}", f"=DATA!{col}{d_row}", bg=bg, fg="008000")
        if ci >= 4: ws_c[f"{get_column_letter(ci)}{r}"].number_format = '$#,##0.00;($#,##0.00);"-"'
    # Residual estimate (monthly × pct)
    fml_cell(ws_c, f"H{r}",
        f"=IFERROR(DATA!E{d_row}*DATA!G{d_row},0)",
        fmt='$#,##0.00;($#,##0.00);"-"', bg=bg, fg="008000")

RATE_END = RATE_START + len(PRODUCTS) - 1
h(ws_c, RATE_END + 1, 8)

# Deal log
LOG_HDR = RATE_END + 2
LOG_COL = LOG_HDR + 1
LOG_START = LOG_COL + 1
DEAL_ROWS = 20
LOG_END = LOG_START + DEAL_ROWS - 1

sec_hdr(ws_c, LOG_HDR, 1, 9, "REGISTRO DE CIERRES DEL MES", bg=C["blue2"])
h(ws_c, LOG_HDR, 24)

log_h = ["Fecha", "Cliente/Negocio", "product_id\n(desde DATA)", "Plan Elegido",
         "¿Pago Anual?\n(SÍ/NO)", "Instalación\nCobrada ($)", "Comisión\nInicial ($)",
         "Residual\nMensual ($)", "Notas"]
for ci, lh in enumerate(log_h, 1):
    cc = ws_c.cell(row=LOG_COL, column=ci)
    cc.value = lh; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["red"]); cc.alignment = align(wrap=True); cc.border = border()
h(ws_c, LOG_COL, 36)

for i in range(DEAL_ROWS):
    r = LOG_START + i
    h(ws_c, r, 22)
    inp_cell(ws_c, f"A{r}", fmt="DD/MM/YYYY")   # date
    inp_cell(ws_c, f"B{r}")                       # client name
    inp_cell(ws_c, f"C{r}")                       # product_id
    inp_cell(ws_c, f"D{r}")                       # plan
    inp_cell(ws_c, f"E{r}", "NO")                 # annual?
    inp_cell(ws_c, f"F{r}", fmt='$#,##0.00')      # setup collected
    inp_cell(ws_c, f"G{r}", fmt='$#,##0.00')      # upfront commission
    inp_cell(ws_c, f"H{r}", fmt='$#,##0.00')      # monthly residual
    inp_cell(ws_c, f"I{r}")                        # notes

h(ws_c, LOG_END + 1, 8)

# Earnings summary
EARN_HDR = LOG_END + 2
sec_hdr(ws_c, EARN_HDR, 1, 9, "RESUMEN DE GANANCIAS DEL MES", bg=C["navy"])
h(ws_c, EARN_HDR, 26)

earn_rows = [
    (EARN_HDR+1, "Total Comisiones Iniciales (Upfront):",
     f"=IFERROR(SUM(G{LOG_START}:G{LOG_END}),0)", C["light"], "008000"),
    (EARN_HDR+2, "Total Ingreso Residual Mensual:",
     f"=IFERROR(SUM(H{LOG_START}:H{LOG_END}),0)", C["white"], "008000"),
    (EARN_HDR+3, "Número de Cierres Este Mes:",
     f"=COUNTA(B{LOG_START}:B{LOG_END})", C["light"], "000000"),
    (EARN_HDR+4, "GANANCIA TOTAL DEL MES (Sin Bonos):",
     f"=IFERROR(SUM(G{LOG_START}:G{LOG_END})+SUM(H{LOG_START}:H{LOG_END}),0)",
     C["navy"], "FFFFFF"),
]
for row_n, lbl_t, fml, bg, fgc in earn_rows:
    h(ws_c, row_n, 24)
    ws_c.merge_cells(f"A{row_n}:F{row_n}")
    lbl_cell(ws_c, f"A{row_n}", lbl_t, bold=True, halign="right", bg=bg,
             fg="FFFFFF" if bg == C["navy"] else "000000")
    ws_c.merge_cells(f"G{row_n}:I{row_n}")
    fmt = '$#,##0.00' if row_n != EARN_HDR+3 else '#,##0'
    fml_cell(ws_c, f"G{row_n}", fml, fmt=fmt, bold=True, fg=fgc, bg=bg)

GRAND_ROW_C = EARN_HDR + 4   # keep reference for sheet 4

ws_c.freeze_panes = "A11"

# ════════════════════════════════════════════════════════════════════
#  SHEET 4: BONUS TRACKER
# ════════════════════════════════════════════════════════════════════
ws_b = wb.create_sheet("Resumen de Bonos")
ws_b.sheet_view.showGridLines = False

for col, width in zip("ABCDEF", [22,22,18,18,18,4]):
    w(ws_b, col, width)

ws_b.merge_cells("A1:E1")
hdr_cell(ws_b, "A1", "TODO DIRECTO — RASTREADOR DE BONOS POR VOLUMEN", bg=C["navy"], size=13)
h(ws_b, 1, 34)
ws_b.merge_cells("A2:E2")
hdr_cell(ws_b, "A2", "Objetivos de bonos vienen de la hoja DATA. Ingresa cierres semanales abajo.", bg=C["blue2"], size=9, fg="AABBCC")
h(ws_b, 2, 18); h(ws_b, 3, 8)

# Bonus structure (from DATA)
sec_hdr(ws_b, 4, 1, 5, "ESTRUCTURA DE BONOS (Referencia desde DATA)", bg=C["blue"])
h(ws_b, 4, 24)

for ci, bh in enumerate(["Tipo de Bono","Meta de Cierres","Premio ($)","Período"], 1):
    cc = ws_b.cell(row=5, column=ci)
    cc.value = bh; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["red"]); cc.alignment = align(); cc.border = border()
h(ws_b, 5, 24)

BONUS_DATA_START = DATA_END + 3   # bonus rows in DATA sheet
bonus_periods = ["Por semana", "Por mes", "Por trimestre (3 meses)"]
for i in range(3):
    r = 6 + i
    d_row = BONUS_DATA_START + i
    bg = C["light"] if i % 2 == 0 else C["white"]
    h(ws_b, r, 22)
    fml_cell(ws_b, f"A{r}", f"=DATA!B{d_row}", bg=bg, fg="008000")  # name
    fml_cell(ws_b, f"B{r}", f"=DATA!C{d_row}", bg=bg, fg="008000")  # target
    fml_cell(ws_b, f"C{r}", f"=DATA!D{d_row}", fmt='$#,##0.00', bg=bg, fg="008000")  # prize
    lbl_cell(ws_b, f"D{r}", bonus_periods[i], bg=bg)

h(ws_b, 9, 8)

# Weekly tracker
sec_hdr(ws_b, 10, 1, 5, "RASTREADOR SEMANAL DEL MES", bg=C["blue2"])
h(ws_b, 10, 24)

wk_hdrs = ["Semana","Cierres Logrados\n(Ingresa aquí)","¿Bono Semanal?","Bono Ganado ($)","Acumulado ($)"]
for ci, wh in enumerate(wk_hdrs, 1):
    cc = ws_b.cell(row=11, column=ci)
    cc.value = wh; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["red"]); cc.alignment = align(wrap=True); cc.border = border()
h(ws_b, 11, 36)

for i, wk in enumerate(["Semana 1","Semana 2","Semana 3","Semana 4"]):
    r = 12 + i
    bg = C["light"] if i % 2 == 0 else C["white"]
    h(ws_b, r, 22)
    lbl_cell(ws_b, f"A{r}", wk, bold=True, halign="center", bg=bg)
    inp_cell(ws_b, f"B{r}", 0)
    fml_cell(ws_b, f"C{r}", f'=IF(B{r}>=DATA!C{BONUS_DATA_START},"✅ SÍ","❌ NO")', bg=bg)
    fml_cell(ws_b, f"D{r}", f'=IF(B{r}>=DATA!C{BONUS_DATA_START},DATA!D{BONUS_DATA_START},0)',
             fmt='$#,##0.00;($#,##0.00);"-"', bg=bg, fg="008000")
    if r == 12:
        fml_cell(ws_b, f"E{r}", f"=D{r}", fmt='$#,##0.00', bg=bg, fg="008000")
    else:
        fml_cell(ws_b, f"E{r}", f"=E{r-1}+D{r}", fmt='$#,##0.00', bg=bg, fg="008000")

h(ws_b, 16, 8)

# Monthly + quarterly bonus
sec_hdr(ws_b, 17, 1, 5, "BONOS MENSUALES Y RESUMEN TOTAL", bg=C["blue"])
h(ws_b, 17, 24)

monthly_rows = [
    (18, "Total Cierres del Mes (de hoja Comisiones):",
     f"='Calculadora de Comisiones'!G{EARN_HDR+3}", "0", C["light"], "008000"),
    (19, "Meta Mensual de Cierres:",
     f"=DATA!C{BONUS_DATA_START+1}", "0", C["white"], "008000"),
    (20, "¿Bono Mensual Ganado?",
     f"=IF('Calculadora de Comisiones'!G{EARN_HDR+3}>=DATA!C{BONUS_DATA_START+1},"
     f'"✅ SÍ","❌ NO")', None, C["light"], "000000"),
    (21, "Monto Bono Mensual:",
     f"=IF('Calculadora de Comisiones'!G{EARN_HDR+3}>=DATA!C{BONUS_DATA_START+1},"
     f"DATA!D{BONUS_DATA_START+1},0)", '$#,##0.00', C["white"], "008000"),
    (22, "Total Bonos Semanales del Mes:",
     "=E15", '$#,##0.00', C["light"], "008000"),
    (23, "TOTAL BONOS DEL MES:",
     f"=E15+IF('Calculadora de Comisiones'!G{EARN_HDR+3}>=DATA!C{BONUS_DATA_START+1},"
     f"DATA!D{BONUS_DATA_START+1},0)", '$#,##0.00', C["navy"], "FFFFFF"),
]
for row_n, lbl_t, fml, fmt, bg, fgc in monthly_rows:
    h(ws_b, row_n, 22)
    ws_b.merge_cells(f"A{row_n}:C{row_n}")
    lbl_cell(ws_b, f"A{row_n}", lbl_t, bold=(row_n==23), halign="right", bg=bg,
             fg="FFFFFF" if bg == C["navy"] else "000000")
    ws_b.merge_cells(f"D{row_n}:E{row_n}")
    fml_cell(ws_b, f"D{row_n}", fml, fmt=fmt, bold=(row_n==23), fg=fgc, bg=bg)

h(ws_b, 24, 8)

# Grand total
sec_hdr(ws_b, 25, 1, 5, "GANANCIA TOTAL DEL MES (Comisiones + Bonos)", bg=C["navy"])
h(ws_b, 25, 26)

grand_rows = [
    (26, "Comisiones del Mes (sin bonos):",
     f"='Calculadora de Comisiones'!G{EARN_HDR+4}", C["light"], "008000"),
    (27, "Total Bonos del Mes:",
     "=D23", C["white"], "008000"),
    (28, "GANANCIA BRUTA TOTAL DEL MES:",
     f"='Calculadora de Comisiones'!G{EARN_HDR+4}+D23", C["red"], "FFFFFF"),
]
for row_n, lbl_t, fml, bg, fgc in grand_rows:
    h(ws_b, row_n, 26)
    ws_b.merge_cells(f"A{row_n}:C{row_n}")
    lbl_cell(ws_b, f"A{row_n}", lbl_t, bold=(row_n==28), halign="right", bg=bg,
             fg="FFFFFF" if bg in [C["navy"], C["red"]] else "000000")
    ws_b.merge_cells(f"D{row_n}:E{row_n}")
    fml_cell(ws_b, f"D{row_n}", fml, fmt='$#,##0.00', bold=(row_n==28), fg=fgc, bg=bg)

# ════════════════════════════════════════════════════════════════════
#  SHEET 5: WEBAPP EXPORT (flat table for JSON/API)
# ════════════════════════════════════════════════════════════════════
ws_ex = wb.create_sheet("WebApp Data Export")
ws_ex.sheet_view.showGridLines = False

for col, width in zip("ABCDEFGHIJ", [14,22,28,14,14,14,14,14,14,12]):
    w(ws_ex, col, width)

ws_ex.merge_cells("A1:J1")
hdr_cell(ws_ex, "A1",
    "TABLA DE EXPORTACIÓN PARA WEB APP — Copia esta tabla en tu base de datos JSON/SQL para portar la lógica al sitio web.",
    bg=C["navy"], size=11)
h(ws_ex, 1, 36)
ws_ex.merge_cells("A2:J2")
hdr_cell(ws_ex, "A2",
    "Todos los valores se actualizan automáticamente desde la hoja DATA. Exporta/copia como JSON para tu desarrollador.",
    bg=C["blue2"], size=9, fg="AABBCC")
h(ws_ex, 2, 18); h(ws_ex, 3, 8)

# Export column headers
export_cols = ["product_id","service","plan","client_setup","client_monthly",
               "rep_upfront","rep_residual_pct","rep_annual","category","active"]
for ci, ec in enumerate(export_cols, 1):
    cc = ws_ex.cell(row=4, column=ci)
    cc.value = ec; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["blue"]); cc.alignment = align(); cc.border = border()
h(ws_ex, 4, 28)

# Pull all data from DATA sheet via formula references
for i in range(len(PRODUCTS)):
    r = 5 + i
    d_row = DATA_START + i
    h(ws_ex, r, 22)
    bg = C["gray"] if i % 2 == 0 else C["white"]
    data_cols = "ABCDEFGHI"
    for ci, dc in enumerate(data_cols, 1):
        fml_cell(ws_ex, f"{get_column_letter(ci)}{r}", f"=DATA!{dc}{d_row}", bg=bg, fg=C["dkgray"])
    # active
    fml_cell(ws_ex, f"J{r}", f"=DATA!J{d_row}", bg=bg, fg=C["dkgray"])

EX_END = 5 + len(PRODUCTS) - 1
h(ws_ex, EX_END + 1, 8)

# Bonus table export
ws_ex.merge_cells(f"A{EX_END+2}:D{EX_END+2}")
hdr_cell(ws_ex, f"A{EX_END+2}", "TABLA DE BONOS PARA WEB APP", bg=C["blue"])
h(ws_ex, EX_END+2, 22)

for ci, bh in enumerate(["bonus_id","nombre","target_deals","prize_usd"], 1):
    cc = ws_ex.cell(row=EX_END+3, column=ci)
    cc.value = bh; cc.font = ft(bold=True, color="FFFFFF", size=9)
    cc.fill = fill(C["red"]); cc.alignment = align(); cc.border = border()
h(ws_ex, EX_END+3, 22)

for i in range(3):
    r = EX_END + 4 + i
    d_row = BONUS_DATA_START + i
    bg = C["gray"] if i % 2 == 0 else C["white"]
    h(ws_ex, r, 20)
    for ci, dc in enumerate(["A","B","C","D"], 1):
        fml_cell(ws_ex, f"{get_column_letter(ci)}{r}", f"=DATA!{dc}{d_row}", bg=bg, fg=C["dkgray"])

# Dev instructions
INST_R = EX_END + 8
ws_ex.merge_cells(f"A{INST_R}:J{INST_R}")
h(ws_ex, INST_R, 50)
c_inst = ws_ex[f"A{INST_R}"]
c_inst.value = (
    "💻  PARA EL DESARROLLADOR WEB:\n"
    "Todos los IDs (product_id, bonus_id) están en inglés para facilitar la integración con APIs.\n"
    "Porta esta tabla a JSON usando: { product_id, service, plan, client_setup, client_monthly, rep_upfront, rep_residual_pct, rep_annual, category, active }\n"
    "Los campos 'active' y 'category' permiten filtrar servicios por tipo de cliente en la webapp.\n"
    "La lógica del cotizador: si include=true → total_monthly = client_setup + client_monthly; total_annual = client_setup + (client_monthly × 10)\n"
    "La lógica de comisión: upfront = rep_upfront; residual = client_monthly × rep_residual_pct (mensual)"
)
c_inst.font = Font(name="Calibri", size=9, italic=True, color="333333")
c_inst.fill = fill("FFF3CD")
c_inst.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
c_inst.border = border()

# ─── Final settings ───────────────────────────────────────────────
wb.active = ws_q

wb.save(OUT)
print(f"Saved: {OUT}")
