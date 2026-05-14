import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
import os

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "Todo_Directo_Commission_Calculator.xlsx")

wb = openpyxl.Workbook()

# ─── Color palette ───────────────────────────────────────────────
BG_DARK     = "1A1A2E"   # dark navy header
BG_MID      = "16213E"   # section header
BG_ACCENT   = "0F3460"   # sub-header
BG_GOLD     = "E94560"   # accent / highlights (Todo Directo brand red)
BG_GOLD2    = "F5A623"   # warm gold for bonuses
BG_WHITE    = "FFFFFF"
BG_LIGHT    = "F0F4FF"   # alternating row light blue
BG_INPUT    = "EBF5FB"   # input cells bg (blue-tinted)
BG_FORMULA  = "FFFFFF"   # formula cells bg

FG_WHITE    = "FFFFFF"
FG_BLUE     = "0000FF"   # blue text = hard-coded inputs (per skill spec)
FG_BLACK    = "000000"   # black text = formulas
FG_GREEN    = "008000"   # green = cross-sheet links
FG_GOLD2    = "E67E22"   # orange label text

# ─── Border helpers ─────────────────────────────────────────────
thin  = Side(style="thin",   color="C5D8F0")
thick = Side(style="medium", color="0F3460")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def thick_border():
    return Border(left=thick, right=thick, top=thick, bottom=thick)

# ─── Style helpers ───────────────────────────────────────────────
def hdr(ws, cell_ref, text, bg=BG_DARK, fg=FG_WHITE, size=11, bold=True, wrap=False, align="center"):
    c = ws[cell_ref]
    c.value = text
    c.font  = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill  = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

def inp(ws, cell_ref, value=None, fmt=None, comment=None):
    """Blue input cell."""
    c = ws[cell_ref]
    c.value = value
    c.font  = Font(name="Calibri", bold=False, color=FG_BLUE, size=10)
    c.fill  = PatternFill("solid", start_color=BG_INPUT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    if fmt:
        c.number_format = fmt
    return c

def formula_cell(ws, cell_ref, formula, fmt=None, fg=FG_BLACK, bg=BG_FORMULA, bold=False):
    """Black formula cell."""
    c = ws[cell_ref]
    c.value = formula
    c.font  = Font(name="Calibri", bold=bold, color=fg, size=10)
    c.fill  = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
    if fmt:
        c.number_format = fmt
    return c

def label(ws, cell_ref, text, bg=BG_WHITE, fg=FG_BLACK, size=10, bold=False, align="left", wrap=False):
    c = ws[cell_ref]
    c.value = text
    c.font  = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill  = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = thin_border()
    return c

def section_hdr(ws, row, col_start, col_end, text, bg=BG_ACCENT):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value = text
    c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=11)
    c.fill  = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

def col_width(ws, col_letter, w):
    ws.column_dimensions[col_letter].width = w

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h


# ════════════════════════════════════════════════════════════════════
#  SHEET 1 – COTIZADOR DE CLIENTES
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Cotizador de Clientes"
ws1.sheet_view.showGridLines = False

# — Column widths
for col, w in zip("ABCDEFGHI", [18, 30, 18, 18, 18, 18, 18, 18, 18]):
    col_width(ws1, col, w)

# ── TITLE BANNER ──────────────────────────────────────────────────
ws1.merge_cells("A1:I1")
c = ws1["A1"]
c.value = "TODO DIRECTO – COTIZADOR DE CLIENTE"
c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=16)
c.fill  = PatternFill("solid", start_color=BG_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
row_height(ws1, 1, 36)

ws1.merge_cells("A2:I2")
c = ws1["A2"]
c.value = "Herramienta de Cotización para Representantes Comerciales | Todo Directo"
c.font  = Font(name="Calibri", italic=True, color="AABBCC", size=10)
c.fill  = PatternFill("solid", start_color=BG_MID)
c.alignment = Alignment(horizontal="center", vertical="center")
row_height(ws1, 2, 20)

# ── CLIENT INFO ───────────────────────────────────────────────────
row_height(ws1, 3, 8)

section_hdr(ws1, 4, 1, 9, "📋  INFORMACIÓN DEL CLIENTE", BG_MID)
row_height(ws1, 4, 24)

fields = [
    ("A5", "B5", "Nombre del Cliente:",   "C5", "D5"),
    ("A6", "B6", "Nombre del Negocio:",   "C6", "D6"),
    ("A7", "B7", "Tipo de Negocio:",      "C7", "D7"),
    ("A8", "B8", "Teléfono:",             "E8", "F8"),
    ("A8", "B8", None,                    None,  None),
]
row_height(ws1, 5, 22); row_height(ws1, 6, 22); row_height(ws1, 7, 22)

label(ws1, "A5", "Nombre del Cliente:", bold=True, align="right")
ws1.merge_cells("B5:D5"); inp(ws1, "B5", "")
label(ws1, "E5", "Representante:",      bold=True, align="right")
ws1.merge_cells("F5:H5"); inp(ws1, "F5", "")

label(ws1, "A6", "Nombre del Negocio:", bold=True, align="right")
ws1.merge_cells("B6:D6"); inp(ws1, "B6", "")
label(ws1, "E6", "Fecha:",              bold=True, align="right")
ws1.merge_cells("F6:H6"); inp(ws1, "F6", None, fmt="DD/MM/YYYY")

label(ws1, "A7", "Tipo de Negocio:",    bold=True, align="right")
ws1.merge_cells("B7:D7"); inp(ws1, "B7", "")
label(ws1, "E7", "Ciudad:",             bold=True, align="right")
ws1.merge_cells("F7:H7"); inp(ws1, "F7", "")

row_height(ws1, 8, 8)

# ── PRICING HEADER ────────────────────────────────────────────────
section_hdr(ws1, 9, 1, 9, "💼  SERVICIOS Y COTIZACIÓN", BG_ACCENT)
row_height(ws1, 9, 24)

col_headers = [
    "Servicio", "Plan/Paquete", "¿Incluir?\n(SÍ / NO)",
    "Cuota de\nInstalación ($)", "Mensualidad\nCliente ($)",
    "Opción Anual\n(×10 meses)", "Instalación\nCon Desc.", "Total Año\n(Anual)",
    "Total Mes\n(Mensual)"
]
for ci, h in enumerate(col_headers, start=1):
    c = ws1.cell(row=10, column=ci)
    c.value = h
    c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=9)
    c.fill  = PatternFill("solid", start_color=BG_GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
row_height(ws1, 10, 36)

# Define products: (Service_label, Plan_label, setup_default, monthly_default)
products = [
    # AlToque
    ("AlToque App",    "Chivo (Gratis)",    0,    0),
    ("AlToque App",    "Pro",               0,   14.99),
    ("AlToque App",    "Premium",           0,   29.99),
    # Web Dev
    ("Desarrollo Web", "Plan Chivo",       99,   25),
    ("Desarrollo Web", "Plan Pro",        149,   49),
    ("Desarrollo Web", "Plan Premium",    299,   89),
    # Vacation Rentals
    ("Vacation Rentals","Modelo A (Direct)", 300,   0),
    ("Vacation Rentals","Modelo B (Full Mgmt)", 600, 0),
    # Automatización
    ("Automatización", "Básica",             0,  200),
    ("Automatización", "Pro",                0,  400),
    # Chatbot IA
    ("Chatbot IA",    "Básico",              0,  150),
    ("Chatbot IA",    "Pro",                 0,  300),
]

START_ROW = 11
# Alternating row bg
row_bgs = [BG_LIGHT, BG_WHITE]

for i, (svc, plan, setup, monthly) in enumerate(products):
    r = START_ROW + i
    row_height(ws1, r, 22)
    bg = row_bgs[i % 2]

    # A: Service
    label(ws1, f"A{r}", svc, bg=bg, bold=False, align="center")
    # B: Plan
    label(ws1, f"B{r}", plan, bg=bg, bold=False, align="left")
    # C: Include (input)
    inp(ws1, f"C{r}", "NO")
    # D: Setup (input – blue)
    inp(ws1, f"D{r}", setup, fmt='$#,##0.00;($#,##0.00);"-"')
    # E: Monthly (input – blue)
    inp(ws1, f"E{r}", monthly, fmt='$#,##0.00;($#,##0.00);"-"')
    # F: Annual option price (10 months × monthly, formula)
    formula_cell(ws1, f"F{r}",
        f'=IF(E{r}>0, E{r}*10, "-")',
        fmt='$#,##0.00;($#,##0.00);"-"')
    # G: Setup with annual discount (20% off setup)
    formula_cell(ws1, f"G{r}",
        f'=IF(UPPER(C{r})="SÍ", D{r}*0.8, "-")',
        fmt='$#,##0.00;($#,##0.00);"-"')
    # H: Total Year (Annual): discounted setup + 10×monthly
    formula_cell(ws1, f"H{r}",
        f'=IF(UPPER(C{r})="SÍ", IF(ISNUMBER(D{r}*0.8+E{r}*10), D{r}*0.8+E{r}*10, "-"), "-")',
        fmt='$#,##0.00;($#,##0.00);"-"')
    # I: Total Month (Monthly): setup + 1×monthly
    formula_cell(ws1, f"I{r}",
        f'=IF(UPPER(C{r})="SÍ", D{r}+E{r}, "-")',
        fmt='$#,##0.00;($#,##0.00);"-"')

END_ROW = START_ROW + len(products) - 1
SUMMARY_ROW = END_ROW + 2

row_height(ws1, END_ROW + 1, 8)

# ── TOTALS SECTION ────────────────────────────────────────────────
section_hdr(ws1, SUMMARY_ROW, 1, 9, "💰  RESUMEN DE COTIZACIÓN", BG_MID)
row_height(ws1, SUMMARY_ROW, 24)

TOTAL_SETUP_ROW   = SUMMARY_ROW + 1
TOTAL_MONTHLY_ROW = SUMMARY_ROW + 2
TOTAL_ANNUAL_ROW  = SUMMARY_ROW + 3
SAVINGS_ROW       = SUMMARY_ROW + 4

for tr in [TOTAL_SETUP_ROW, TOTAL_MONTHLY_ROW, TOTAL_ANNUAL_ROW, SAVINGS_ROW]:
    row_height(ws1, tr, 22)

ws1.merge_cells(f"A{TOTAL_SETUP_ROW}:F{TOTAL_SETUP_ROW}")
label(ws1, f"A{TOTAL_SETUP_ROW}", "Total Cuota de Instalación (Mensual):",
      bold=True, align="right", bg=BG_LIGHT)
ws1.merge_cells(f"G{TOTAL_SETUP_ROW}:I{TOTAL_SETUP_ROW}")
formula_cell(ws1, f"G{TOTAL_SETUP_ROW}",
    f"=SUMIF(C{START_ROW}:C{END_ROW},\"SÍ\",D{START_ROW}:D{END_ROW})",
    fmt='$#,##0.00', bold=True, bg=BG_LIGHT)

ws1.merge_cells(f"A{TOTAL_MONTHLY_ROW}:F{TOTAL_MONTHLY_ROW}")
label(ws1, f"A{TOTAL_MONTHLY_ROW}", "Total Mensualidad del Cliente:",
      bold=True, align="right", bg=BG_WHITE)
ws1.merge_cells(f"G{TOTAL_MONTHLY_ROW}:I{TOTAL_MONTHLY_ROW}")
formula_cell(ws1, f"G{TOTAL_MONTHLY_ROW}",
    f"=SUMIF(C{START_ROW}:C{END_ROW},\"SÍ\",E{START_ROW}:E{END_ROW})",
    fmt='$#,##0.00', bold=True, bg=BG_WHITE)

ws1.merge_cells(f"A{TOTAL_ANNUAL_ROW}:F{TOTAL_ANNUAL_ROW}")
label(ws1, f"A{TOTAL_ANNUAL_ROW}", "Total Opción Anual (Cliente Paga):",
      bold=True, align="right", bg=BG_LIGHT)
ws1.merge_cells(f"G{TOTAL_ANNUAL_ROW}:I{TOTAL_ANNUAL_ROW}")
formula_cell(ws1, f"G{TOTAL_ANNUAL_ROW}",
    f"=SUMIF(C{START_ROW}:C{END_ROW},\"SÍ\",H{START_ROW}:H{END_ROW})",
    fmt='$#,##0.00', bold=True, bg=BG_LIGHT)

ws1.merge_cells(f"A{SAVINGS_ROW}:F{SAVINGS_ROW}")
label(ws1, f"A{SAVINGS_ROW}", "Ahorro del Cliente (Anual vs Mensual x12):",
      bold=True, align="right", bg=BG_WHITE, fg=FG_GOLD2)
ws1.merge_cells(f"G{SAVINGS_ROW}:I{SAVINGS_ROW}")
formula_cell(ws1, f"G{SAVINGS_ROW}",
    f"=SUMIF(C{START_ROW}:C{END_ROW},\"SÍ\",E{START_ROW}:E{END_ROW})*12"
    f"-G{TOTAL_ANNUAL_ROW}",
    fmt='$#,##0.00', bold=True, fg=FG_GOLD2, bg=BG_WHITE)

# Note row
NOTE_ROW = SAVINGS_ROW + 2
row_height(ws1, NOTE_ROW, 30)
ws1.merge_cells(f"A{NOTE_ROW}:I{NOTE_ROW}")
c = ws1[f"A{NOTE_ROW}"]
c.value = ("⚠️  INSTRUCCIONES: En columna C escribe 'SÍ' para incluir el servicio en la cotización o 'NO' para excluirlo.  "
           "Puedes modificar precios en columnas D y E.  La cotización se actualiza automáticamente.")
c.font  = Font(name="Calibri", italic=True, color="555555", size=9)
c.fill  = PatternFill("solid", start_color="FFF8DC")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = thin_border()


# ════════════════════════════════════════════════════════════════════
#  SHEET 2 – CALCULADORA DE COMISIONES
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Calculadora de Comisiones")
ws2.sheet_view.showGridLines = False

for col, w in zip("ABCDEFGHIJ", [22, 28, 16, 16, 16, 16, 18, 18, 18, 3]):
    col_width(ws2, col, w)

# Title
ws2.merge_cells("A1:I1")
c = ws2["A1"]
c.value = "TODO DIRECTO – CALCULADORA DE COMISIONES DEL REPRESENTANTE"
c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=15)
c.fill  = PatternFill("solid", start_color=BG_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
row_height(ws2, 1, 36)

ws2.merge_cells("A2:I2")
c = ws2["A2"]
c.value = "Actualiza la Columna C del 'Cotizador' primero. Luego ingresa los datos del cierre aquí."
c.font  = Font(name="Calibri", italic=True, color="AABBCC", size=10)
c.fill  = PatternFill("solid", start_color=BG_MID)
c.alignment = Alignment(horizontal="center", vertical="center")
row_height(ws2, 2, 20)

row_height(ws2, 3, 8)

# ── Rep Info ──────────────────────────────────────────────────────
section_hdr(ws2, 4, 1, 9, "👤  INFORMACIÓN DEL REPRESENTANTE", BG_MID)
row_height(ws2, 4, 24)
row_height(ws2, 5, 22)
row_height(ws2, 6, 22)

label(ws2, "A5", "Nombre del Representante:", bold=True, align="right")
ws2.merge_cells("B5:D5"); inp(ws2, "B5", "")
label(ws2, "E5", "Mes de Seguimiento:", bold=True, align="right")
ws2.merge_cells("F5:H5"); inp(ws2, "F5", None, fmt="MMMM YYYY")

label(ws2, "A6", "Meta de Cierres del Mes:", bold=True, align="right")
ws2.merge_cells("B6:D6"); inp(ws2, "B6", 20)
label(ws2, "E6", "Cierres Logrados Hasta Hoy:", bold=True, align="right")
ws2.merge_cells("F6:H6"); inp(ws2, "F6", 0)

row_height(ws2, 7, 8)

# ── Commission Rates (inputs) ─────────────────────────────────────
section_hdr(ws2, 8, 1, 9, "⚙️  TASAS DE COMISIÓN (No modificar a menos que el supervisor lo autorice)", BG_ACCENT)
row_height(ws2, 8, 24)

rate_headers = ["Producto", "Fuente de Comisión", "Tasa / Monto Fijo"]
for ci, h in enumerate(rate_headers, 1):
    c = ws2.cell(row=9, column=ci)
    c.value = h
    c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=9)
    c.fill  = PatternFill("solid", start_color=BG_GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    row_height(ws2, 9, 22)

rate_rows = [
    ("AlToque Pro",            "150% del primer mes",             22.49),
    ("AlToque Premium",        "150% del primer mes",             44.98),
    ("AlToque Pro (Anual)",    "Pago anticipado anual",           37.50),
    ("AlToque Premium (Anual)","Pago anticipado anual",           75.00),
    ("Web Plan Chivo",         "40% de instalación + 10%/mes",    39.60),
    ("Web Plan Pro",           "40% de instalación + 10%/mes",    59.60),
    ("Web Plan Premium",       "40% de instalación + 10%/mes",   119.60),
    ("Vacation Rentals A",     "Comisión fija",                  300.00),
    ("Vacation Rentals B",     "Comisión fija",                  600.00),
    ("Automatización Básica",  "Primer mes + 15% residual",       100.00),
    ("Automatización Pro",     "Primer mes + 15% residual",       200.00),
    ("Chatbot IA Básico",      "Primer mes + 15% residual",        75.00),
    ("Chatbot IA Pro",         "Primer mes + 15% residual",       150.00),
]

RATE_START = 10
for i, (prod, src, rate) in enumerate(rate_rows):
    r = RATE_START + i
    row_height(ws2, r, 20)
    bg = row_bgs[i % 2]
    label(ws2, f"A{r}", prod, bg=bg, align="left")
    label(ws2, f"B{r}", src, bg=bg, align="left", fg="444444")
    inp(ws2, f"C{r}", rate, fmt='$#,##0.00;($#,##0.00);"-"')

RATE_END = RATE_START + len(rate_rows) - 1
row_height(ws2, RATE_END + 1, 8)

# ── Deal Log ──────────────────────────────────────────────────────
LOG_HDR_ROW  = RATE_END + 2
LOG_COL_ROW  = LOG_HDR_ROW + 1
LOG_DATA_ROW = LOG_COL_ROW + 1

section_hdr(ws2, LOG_HDR_ROW, 1, 9, "📝  REGISTRO DE CIERRES DEL MES", BG_MID)
row_height(ws2, LOG_HDR_ROW, 24)

log_headers = [
    "Fecha", "Cliente / Negocio", "Producto/Plan",
    "¿Pago Anual?\n(SÍ/NO)", "Instalación\nCobrada ($)",
    "Mensualidad\nCliente ($)", "Comisión\nInicial ($)",
    "Residual\nMensual ($)", "Notas"
]
for ci, h in enumerate(log_headers, 1):
    c = ws2.cell(row=LOG_COL_ROW, column=ci)
    c.value = h
    c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=9)
    c.fill  = PatternFill("solid", start_color=BG_GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
row_height(ws2, LOG_COL_ROW, 36)

# 20 blank deal rows
DEAL_ROWS = 20
for i in range(DEAL_ROWS):
    r = LOG_DATA_ROW + i
    bg = row_bgs[i % 2]
    row_height(ws2, r, 20)
    inp(ws2, f"A{r}", None, fmt="DD/MM/YYYY")
    inp(ws2, f"B{r}", None)
    inp(ws2, f"C{r}", None)
    inp(ws2, f"D{r}", "NO")
    inp(ws2, f"E{r}", None, fmt='$#,##0.00;($#,##0.00);"-"')
    inp(ws2, f"F{r}", None, fmt='$#,##0.00;($#,##0.00);"-"')
    inp(ws2, f"G{r}", None, fmt='$#,##0.00;($#,##0.00);"-"')
    inp(ws2, f"H{r}", None, fmt='$#,##0.00;($#,##0.00);"-"')
    inp(ws2, f"I{r}", None)

LOG_END_ROW = LOG_DATA_ROW + DEAL_ROWS - 1

# ── Commission Summary ────────────────────────────────────────────
SUM_HDR = LOG_END_ROW + 2
section_hdr(ws2, SUM_HDR, 1, 9, "💵  RESUMEN DE GANANCIAS DEL MES", BG_DARK)
row_height(ws2, SUM_HDR, 26)

summary_rows = [
    (SUM_HDR+1, "Total Comisiones Iniciales (Cierres):", f"=IFERROR(SUM(G{LOG_DATA_ROW}:G{LOG_END_ROW}),0)", BG_LIGHT),
    (SUM_HDR+2, "Total Ingreso Residual Mensual:",        f"=IFERROR(SUM(H{LOG_DATA_ROW}:H{LOG_END_ROW}),0)", BG_WHITE),
    (SUM_HDR+3, "Número de Cierres Este Mes:",            f"=COUNTA(B{LOG_DATA_ROW}:B{LOG_END_ROW})",          BG_LIGHT),
    (SUM_HDR+4, "Progreso hacia Meta (%):",               f"=IFERROR(COUNTA(B{LOG_DATA_ROW}:B{LOG_END_ROW})/B6,0)", BG_WHITE),
]
for r, lbl_txt, formula, bg in summary_rows:
    row_height(ws2, r, 22)
    ws2.merge_cells(f"A{r}:E{r}")
    label(ws2, f"A{r}", lbl_txt, bold=True, align="right", bg=bg)
    ws2.merge_cells(f"F{r}:I{r}")
    fmt = '$#,##0.00' if r in [SUM_HDR+1, SUM_HDR+2] else ('0' if r == SUM_HDR+3 else '0.0%')
    formula_cell(ws2, f"F{r}", formula, fmt=fmt, bold=True, bg=bg,
                 fg="008000" if r in [SUM_HDR+1, SUM_HDR+2] else FG_BLACK)

# Pre-Total row (initial + residual)
GRAND_ROW = SUM_HDR + 5
row_height(ws2, GRAND_ROW, 26)
ws2.merge_cells(f"A{GRAND_ROW}:E{GRAND_ROW}")
label(ws2, f"A{GRAND_ROW}", "GANANCIA TOTAL DEL MES (Sin Bonos):",
      bold=True, align="right", bg=BG_ACCENT, fg=FG_WHITE)
ws2.merge_cells(f"F{GRAND_ROW}:I{GRAND_ROW}")
formula_cell(ws2, f"F{GRAND_ROW}",
    f"=IFERROR(SUM(G{LOG_DATA_ROW}:G{LOG_END_ROW})+SUM(H{LOG_DATA_ROW}:H{LOG_END_ROW}),0)",
    fmt='$#,##0.00', bold=True, fg=FG_WHITE, bg=BG_ACCENT)


# ════════════════════════════════════════════════════════════════════
#  SHEET 3 – RESUMEN DE BONOS
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Resumen de Bonos")
ws3.sheet_view.showGridLines = False

for col, w in zip("ABCDEFG", [24, 22, 22, 18, 18, 18, 4]):
    col_width(ws3, col, w)

# Title
ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "TODO DIRECTO – RASTREADOR DE BONOS POR VOLUMEN"
c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=15)
c.fill  = PatternFill("solid", start_color=BG_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
row_height(ws3, 1, 36)

ws3.merge_cells("A2:F2")
c = ws3["A2"]
c.value = "Los bonos se calculan automáticamente con base en tus cierres registrados en el mes."
c.font  = Font(name="Calibri", italic=True, color="AABBCC", size=10)
c.fill  = PatternFill("solid", start_color=BG_MID)
c.alignment = Alignment(horizontal="center", vertical="center")
row_height(ws3, 2, 20)

row_height(ws3, 3, 8)

# ── Bonus Structure Table ─────────────────────────────────────────
section_hdr(ws3, 4, 1, 6, "🏆  ESTRUCTURA DE BONOS (Solo lectura — aprobado por Todo Directo)", BG_MID)
row_height(ws3, 4, 24)

bonus_col_hdrs = ["Tipo de Bono", "Objetivo de Cierres", "Premio ($)", "Período"]
for ci, h in enumerate(bonus_col_hdrs, 1):
    c = ws3.cell(row=5, column=ci)
    c.value = h
    c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=10)
    c.fill  = PatternFill("solid", start_color=BG_GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
row_height(ws3, 5, 24)

bonus_structure = [
    ("Bono Semanal",     5,  50,  "Por semana"),
    ("Bono Mensual",    20, 200,  "Por mes"),
    ("Bono Trimestral", 60, 500,  "Por trimestre (3 meses)"),
]
BONUS_RATES = {6: 5, 7: 20, 8: 60}   # row → target deals
BONUS_AMT   = {6: 50, 7: 200, 8: 500}

for i, (btype, target, prize, period) in enumerate(bonus_structure):
    r = 6 + i
    bg = row_bgs[i % 2]
    row_height(ws3, r, 22)
    label(ws3, f"A{r}", btype,  bg=bg, bold=True, align="center")
    inp(ws3,   f"B{r}", target)
    inp(ws3,   f"C{r}", prize,  fmt='$#,##0.00')
    label(ws3, f"D{r}", period, bg=bg, align="center")

row_height(ws3, 9, 8)

# ── Monthly Tracker ───────────────────────────────────────────────
section_hdr(ws3, 10, 1, 6, "📅  RASTREADOR MENSUAL DE BONOS", BG_ACCENT)
row_height(ws3, 10, 26)

tracker_hdrs = ["Semana / Período", "Cierres Logrados\n(Ingresa aquí)", "Meta de Cierres",
                "¿Bono Ganado?", "Monto del Bono ($)", "Bono Acumulado ($)"]
for ci, h in enumerate(tracker_hdrs, 1):
    c = ws3.cell(row=11, column=ci)
    c.value = h
    c.font  = Font(name="Calibri", bold=True, color=FG_WHITE, size=9)
    c.fill  = PatternFill("solid", start_color=BG_GOLD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
row_height(ws3, 11, 36)

weeks = ["Semana 1", "Semana 2", "Semana 3", "Semana 4"]
for i, wk in enumerate(weeks):
    r = 12 + i
    bg = row_bgs[i % 2]
    row_height(ws3, r, 22)
    label(ws3, f"A{r}", wk, bg=bg, align="center", bold=True)
    inp(ws3,   f"B{r}", 0)
    # C: Weekly target
    label(ws3, f"C{r}", "=B6", bg=bg, align="center")   # refers to bonus table B6
    ws3[f"C{r}"].value = "=B6"
    ws3[f"C{r}"].font  = Font(name="Calibri", color=FG_GREEN, size=10)
    ws3[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")
    ws3[f"C{r}"].fill  = PatternFill("solid", start_color=bg)
    ws3[f"C{r}"].border = thin_border()
    # D: ¿Bono ganado?
    formula_cell(ws3, f"D{r}", f'=IF(B{r}>=B6,"✅ SÍ","❌ NO")', bg=bg)
    # E: Bonus amount
    formula_cell(ws3, f"E{r}", f'=IF(B{r}>=B6,C6,0)', fmt='$#,##0.00;($#,##0.00);"-"', bg=bg)
    # F: Running total (accumulated)
    if r == 12:
        formula_cell(ws3, f"F{r}", f'=E{r}', fmt='$#,##0.00', bg=bg, fg=FG_GREEN)
    else:
        formula_cell(ws3, f"F{r}", f'=F{r-1}+E{r}', fmt='$#,##0.00', bg=bg, fg=FG_GREEN)

row_height(ws3, 16, 8)

# Monthly bonus row
section_hdr(ws3, 17, 1, 6, "📊  BONO MENSUAL (Calculado automáticamente)", BG_MID)
row_height(ws3, 17, 24)

monthly_rows = [
    (18, "Total Cierres del Mes:",          f"='Calculadora de Comisiones'!{chr(70)}{SUM_HDR+3}", FG_GREEN, BG_LIGHT),
    (19, "Meta Mensual de Cierres:",        "=B7",                                               FG_GREEN, BG_WHITE),
    (20, "¿Bono Mensual Ganado?",           f"=IF('Calculadora de Comisiones'!F{SUM_HDR+3}>=B7,\"✅ SÍ\",\"❌ NO\")", FG_BLACK, BG_LIGHT),
    (21, "Monto Bono Mensual:",             f"=IF('Calculadora de Comisiones'!F{SUM_HDR+3}>=B7,C7,0)", FG_GREEN, BG_WHITE),
    (22, "Total Bonos Semanales del Mes:",  "=F15",                                              FG_GREEN, BG_LIGHT),
    (23, "TOTAL BONOS DEL MES:",            "=F15+IF('Calculadora de Comisiones'!"
                                            f"F{SUM_HDR+3}>=B7,C7,0)",                          FG_WHITE, BG_DARK),
]

for r, lbl_txt, formula, fg, bg in monthly_rows:
    row_height(ws3, r, 22)
    ws3.merge_cells(f"A{r}:C{r}")
    label(ws3, f"A{r}", lbl_txt, bold=(r==23), align="right", bg=bg,
          fg=FG_WHITE if r==23 else FG_BLACK)
    ws3.merge_cells(f"D{r}:F{r}")
    fmt = '$#,##0.00' if r in [21, 22, 23] else '0'
    if r in [20]:
        formula_cell(ws3, f"D{r}", formula, fg=fg, bg=bg, bold=(r==23))
    else:
        formula_cell(ws3, f"D{r}", formula, fmt=fmt, fg=fg, bg=bg, bold=(r==23))

row_height(ws3, 24, 8)

# Grand Total summary
GRAND_SUM_HDR = 25
section_hdr(ws3, GRAND_SUM_HDR, 1, 6, "🏆  RESUMEN TOTAL DE GANANCIAS DEL MES", BG_DARK)
row_height(ws3, GRAND_SUM_HDR, 26)

grand_rows = [
    (GRAND_SUM_HDR+1, "Ganancias por Comisiones (sin bonos):",
     f"='Calculadora de Comisiones'!F{GRAND_ROW}", '$#,##0.00', BG_LIGHT, FG_GREEN),
    (GRAND_SUM_HDR+2, "Total Bonos del Mes:",
     "=D23", '$#,##0.00', BG_WHITE, FG_GREEN),
    (GRAND_SUM_HDR+3, "💰 GANANCIA TOTAL DEL MES:",
     f"='Calculadora de Comisiones'!F{GRAND_ROW}+D23",
     '$#,##0.00', BG_ACCENT, FG_WHITE),
]

for r, lbl_txt, formula, fmt, bg, fg in grand_rows:
    row_height(ws3, r, 24)
    ws3.merge_cells(f"A{r}:C{r}")
    label(ws3, f"A{r}", lbl_txt, bold=True, align="right", bg=bg,
          fg=FG_WHITE if bg == BG_ACCENT else FG_BLACK)
    ws3.merge_cells(f"D{r}:F{r}")
    formula_cell(ws3, f"D{r}", formula, fmt=fmt, fg=fg, bg=bg,
                 bold=(r == GRAND_SUM_HDR+3))


# ─── Freeze panes & active tab ───────────────────────────────────
ws1.freeze_panes = "A11"
ws2.freeze_panes = "A9"
ws3.freeze_panes = "A6"

wb.active = ws1

wb.save(OUTPUT_PATH)
print(f"Saved: {OUTPUT_PATH}")
